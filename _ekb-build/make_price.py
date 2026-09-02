# -*- coding: utf-8 -*-
"""Сборка прайс-листа /price/price-fanline.xlsx.

Запуск:  python3 _ekb-build/make_price.py

Раньше файл лежал в репозитории готовым и жил своей жизнью: в нём осталась
колонка «цена за мешок», которой больше нет на сайте, и обещание самовывоза,
которого у нас нет вовсе. Теперь он собирается из тех же данных, что и
страницы (prices.py и site_config.py), поэтому разойтись они не могут.
"""
import os
import sys
from datetime import date

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(HERE, "data"))
ROOT = os.path.dirname(HERE)

from site_config import SITE          # noqa: E402
from prices import PRICES             # noqa: E402
from cities import CITIES, BASE_KM    # noqa: E402

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Порядок и примечания: как в каталоге на сайте, чтобы прайс читался
# рядом со страницами, а не как отдельный документ.
ROWS = [
    ("Чернозём",          "под грядки, газон, теплицу"),
    ("Перегной",          "перепревший, под посадку"),
    ("Навоз коровий",     "свежий и перепревший"),
    ("Навоз конский",     "под тёплые грядки"),
    ("Торф",              "низинный и верховой"),
    ("Плодородный грунт", "смесь под газон и отсыпку"),
    ("Торфогрунт",        "готовая смесь под рассаду"),
]

INK = "1F3B2C"
ACCENT = "E8F1EA"
THIN = Side(style="thin", color="C9D8CD")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def legs():
    return 2 if SITE["km_round_trip"] else 1


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "Прайс"
    ws.sheet_view.showGridLines = False

    def put(row, col, value, bold=False, size=11, fill=None, wrap=False, align="left"):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name="Calibri", size=size, bold=bold, color=INK)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        if fill:
            c.fill = PatternFill("solid", fgColor=fill)
        return c

    r = 1
    put(r, 1, "Доставка грунта по Екатеринбургу и области", bold=True, size=16); r += 1
    put(r, 1, f"Прайс-лист от {date.today().strftime('%d.%m.%Y')}", size=10); r += 1
    put(r, 1, f"Связь: мессенджер MAX  ·  Почта: {SITE['contact_email']}", size=10); r += 1
    put(r, 1, f"{SITE['hours']}. {SITE['payment']}", size=10); r += 2

    # Условие, из-за которого прайс и переписан: минимум и отсутствие фасовки
    # стоят выше таблицы, а не сноской под ней.
    put(r, 1, f"Минимальный заказ {SITE['min_volume']}: {SITE['min_volume_note']}. "
              f"Возим только навалом, фасовки в мешках нет.", bold=True, size=11); r += 2

    put(r, 1, "Материал", bold=True, fill=ACCENT)
    put(r, 2, "Цена за м³, ₽", bold=True, fill=ACCENT, align="center")
    put(r, 3, "Примечание", bold=True, fill=ACCENT)
    for col in range(1, 4):
        ws.cell(row=r, column=col).border = BOX
    r += 1

    for name, note in ROWS:
        price = PRICES.get(name, {}).get("m3")
        put(r, 1, name)
        put(r, 2, f"от {price}" if price else "по запросу", align="center")
        put(r, 3, note)
        for col in range(1, 4):
            ws.cell(row=r, column=col).border = BOX
        r += 1

    r += 1
    put(r, 1, "Доставка", bold=True, size=13); r += 1
    put(r, 1, f"от {SITE['km_price']} ₽ за километр от базы"
              + (", рейс считается туда и обратно" if SITE["km_round_trip"] else "")
              + ", с подачей машины", size=10); r += 1
    put(r, 1, "Базы разные: земля, торф и торфогрунт грузятся в Курганово "
              "(Полевской тракт), перегной и навоз — в Садовом. Плечо в таблице "
              "ниже дано для земли; по органике считаем от Садового.", size=10); r += 1
    put(r, 1, "Стоимость рейса не зависит от загрузки кузова, поэтому чем больше объём, "
              "тем дешевле выходит кубометр.", size=10); r += 2

    put(r, 1, "Куда", bold=True, fill=ACCENT)
    put(r, 2, "Плечо, км", bold=True, fill=ACCENT, align="center")
    put(r, 3, "Рейс, ₽", bold=True, fill=ACCENT, align="center")
    for col in range(1, 4):
        ws.cell(row=r, column=col).border = BOX
    r += 1

    ZEMLYA = "kurganovo"
    for key in sorted(BASE_KM, key=lambda k: BASE_KM[k][ZEMLYA]):
        km = BASE_KM[key][ZEMLYA]
        put(r, 1, CITIES[key]["name"])
        put(r, 2, km, align="center")
        put(r, 3, km * SITE["km_price"] * legs() + SITE.get("order_fee", 0), align="center")
        for col in range(1, 4):
            ws.cell(row=r, column=col).border = BOX
        r += 1

    r += 1
    put(r, 1, "Цены на материал ориентировочные, «от». Точную стоимость с доставкой "
              "называем по заявке в MAX, когда знаем объём, адрес и подъезд.", size=10)

    for col, width in ((1, 34), (2, 16), (3, 38)):
        ws.column_dimensions[get_column_letter(col)].width = width

    out = os.path.join(ROOT, "price", "price-fanline.xlsx")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    print("Готово:", out)


if __name__ == "__main__":
    build()
